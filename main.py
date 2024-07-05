import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import re
import smtplib

# Carregar o arquivo Excel
excel = load_workbook('Relacao_Produtos_e_Clientes_2024.xlsx')
planilha = excel.active

# Listas para armazenar os dados
lista_produtos = []
lista_valor = []
lista_regiao = []
lista_equipe = []
lista_cliente = []
lista_metPagamento = []
lista_desconto = []

# Função para verificar se o valor a partir da segunda letra é numérico
def verificar_numero(valor):
    try:
        float(valor[1:])
        return True
    except (ValueError, TypeError):
        return False

# Dicionário de substituição para métodos de pagamento
def verificar_pagamento(metPagamento):
    substituicao = {
        'cartao_credito': ['Cartão de Crédito', 'Cred.'], 
        'cartao_debito': ['Cartão de Débito'],
        'transf_bancaria': ['Transferência Bancária', 'Tran. Bancária'],
        'dinheiro': ['Dinheiro'],
        'cheque': ['Cheque']
    }
    for key, values in substituicao.items():
        if metPagamento in values:
            return key
    return metPagamento

# Iterar sobre as linhas da planilha e preencher as listas
for row in planilha.iter_rows(min_row=2, values_only=True):
    produto = row[1]
    valor = row[2]
    regiao = row[3]
    equipe = row[4]
    cliente = row[5]
    metPagamento = row[6]
    desconto = row[7]

    # Adicionar produto à lista de produtos
    lista_produtos.append(produto)

    # Verificar se o valor é numérico e adicionar à lista de valores
    if verificar_numero(valor):
        valor_novo = re.sub(r'[$]', '', valor)  # Remover símbolo $
        lista_valor.append(float(valor_novo))
    else:
        lista_valor.append(None)

    # Adicionar região à lista de regiões
    lista_regiao.append(regiao)

    # Adicionar equipe à lista de equipes
    lista_equipe.append(equipe)

    # Adicionar cliente à lista de clientes
    lista_cliente.append(cliente)

    # Substituir o método de pagamento e adicionar à lista método pagamento
    metodo_pagamento_corrigido = verificar_pagamento(metPagamento)
    lista_metPagamento.append(metodo_pagamento_corrigido)

    # Verificar se o desconto é positivo e adicionar à lista de descontos
    if desconto >= 0:
        lista_desconto.append(desconto)
    else:
        lista_desconto.append(None)

# Criar DataFrame
df = pd.DataFrame({
    'Produto': lista_produtos,
    'Valor da Venda': lista_valor,
    'Região': lista_regiao,
    'Equipe de Venda': lista_equipe,
    'Cliente': lista_cliente,
    'Método de Pagamento': lista_metPagamento,
    'Desconto': lista_desconto
})

# Remover linhas onde Valor da Venda ou Desconto é None
df = df.dropna(subset=['Valor da Venda', 'Desconto'])

# Agrupar por Produto e Região e calcular a média do Valor da Venda
df_grupo1 = df.groupby(['Produto', 'Região'])['Valor da Venda'].mean().unstack()

# Plotar o gráfico de barras agrupadas por região
df_grupo1.plot(kind='bar', figsize=(12, 6), width=(0.8))
plt.xlabel('Produto')
plt.ylabel('Valor da Venda (Média)')
plt.title('Valor Médio das Vendas por Produto e Região')
plt.xticks(rotation=30)
plt.legend(title='Região', bbox_to_anchor=(1.05, 1), loc='upper left')
plt.tight_layout()

# Mostrar o gráfico
plt.show()

# Agrupar por Produto e Equipe de Venda e calcular a média do Valor da Venda
df_grupo2 = df.groupby(['Produto', 'Equipe de Venda'])['Valor da Venda'].mean().unstack()

# Plotar o gráfico de barras agrupadas por equipe
df_grupo2.plot(kind='bar', figsize=(12, 6), width=(0.8))
plt.xlabel('Produto')
plt.ylabel('Valor da Venda (Média)')
plt.title('Valor Médio das Vendas por Produto e Equipe')
plt.xticks(rotation=30)
plt.legend(title='Equipe', bbox_to_anchor=(1.05, 1), loc='upper left')
plt.tight_layout()

# Mostrar o gráfico
plt.show()